import openpyxl
import os

class ExcelConnector:
    """Connector for reading and writing Excel files using openpyxl."""
    
    def read(self, config, fs=None):
        """Read data from Excel file."""
        file_path = config.get('filePath')
        sheet_name = config.get('sheetName', 'Sheet1')
        has_header = config.get('hasHeader', True)
        
        if not file_path:
             raise Exception('Excel file path is required')

        if fs:
            if not fs.exists(file_path):
                 raise Exception(f'Excel file not found: {file_path}')
            # Open as binary
            file_obj = fs.open(file_path, 'rb')
        else:
            if not os.path.exists(file_path):
                raise Exception(f'Excel file not found: {file_path}')
            file_obj = open(file_path, 'rb')
            
        try:
            wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                sheet_name = wb.sheetnames[0]
                
            ws = wb[sheet_name]
            data = []
            rows = list(ws.rows)
            
            if not rows:
                return []
                
            if has_header:
                headers = [cell.value for cell in rows[0]]
                for row in rows[1:]:
                    record = {}
                    for i, cell in enumerate(row):
                        if i < len(headers):
                            record[str(headers[i])] = cell.value
                    data.append(record)
            else:
                for row in rows:
                    record = {}
                    for i, cell in enumerate(row):
                        record[f'col_{i}'] = cell.value
                    data.append(record)
            
            return data
        finally:
            if file_obj and not file_obj.closed and fs:
                 file_obj.close()
            # Local open() context manager is not used here due to wb loading, so we rely on GC or explicit close if we kept handle
            # But openpyxl might require seekable, which fs.open provides.
            # Ideally we wrap in try/finally to close file_obj if we opened it.
            if hasattr(file_obj, 'close'):
               file_obj.close()

    
    def write(self, data, config, fs=None):
        """Write data to Excel file."""
        file_path = config.get('filePath')
        sheet_name = config.get('sheetName', 'Sheet1')
        
        if not file_path:
            raise Exception('Excel file path is required')
            
        if not data:
            return False
        
        # Create directory
        if fs:
            try:
                fs.makedirs(os.path.dirname(file_path), exist_ok=True)
            except:
                pass
        else:
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        # Write headers
        headers = list(data[0].keys())
        ws.append(headers)
        
        # Write data
        for row in data:
            ws.append([row.get(h) for h in headers])
            
        # Save to file object
        if fs:
            with fs.open(file_path, 'wb') as f:
                wb.save(f)
        else:
            wb.save(file_path)
        
        return True
